"""Extrai os dados brutos das planilhas-fonte para CSV.

Fontes:
- arquivos_apoio/ANP Dados - Bases e Terminais.xlsx -> dados_brutos/anp_instalacoes.csv
- arquivos_apoio/exportação (2).xlsx (Folha1)        -> dados_brutos/bases_autorizadas.csv
- arquivos_apoio/exportação (2).xlsx (Planilha1)     -> dados_brutos/empresas_alias.csv
"""

import csv
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
APOIO_DIR = BASE_DIR / "arquivos_apoio"
OUT_DIR = BASE_DIR / "dados_brutos"

ANP_XLSX = APOIO_DIR / "ANP Dados - Bases e Terminais.xlsx"
BASES_XLSX = APOIO_DIR / "exportação (2).xlsx"


def extract_anp_instalacoes():
    wb = openpyxl.load_workbook(ANP_XLSX, read_only=True, data_only=True)
    ws = wb["Planilha1"]
    header = ["Uf", "Municipio", "NomeEmpresarial", "Segmento", "DetalheInstalacao", "NumTanques", "TancagemTotalM3"]
    out_rows = []
    for row in ws.iter_rows(min_row=5, values_only=True):
        uf, municipio, nome_empresarial = row[0], row[1], row[2]
        if not nome_empresarial or uf in ("(vazio)", "Total Geral"):
            continue
        out_rows.append(list(row))
    _write_csv(OUT_DIR / "anp_instalacoes.csv", header, out_rows)
    return len(out_rows)


def extract_bases_autorizadas():
    wb = openpyxl.load_workbook(BASES_XLSX, read_only=True, data_only=True)
    ws = wb["Folha1"]
    rows = list(ws.iter_rows(values_only=True))
    header = list(rows[0])
    out_rows = [list(r) for r in rows[1:] if r[0]]
    _write_csv(OUT_DIR / "bases_autorizadas.csv", header, out_rows)
    return len(out_rows)


def extract_empresas_alias():
    wb = openpyxl.load_workbook(BASES_XLSX, read_only=True, data_only=True)
    ws = wb["Planilha1"]
    rows = list(ws.iter_rows(values_only=True))
    header = ["RazaoSocialOriginal", "NomeReduzido", "TipoEmpresa"]
    out_rows = [list(r) for r in rows[1:] if r[0]]
    _write_csv(OUT_DIR / "empresas_alias.csv", header, out_rows)
    return len(out_rows)


def _write_csv(path, header, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        writer.writerows(rows)


def main():
    OUT_DIR.mkdir(exist_ok=True)
    n_anp = extract_anp_instalacoes()
    n_bases = extract_bases_autorizadas()
    n_alias = extract_empresas_alias()
    print(f"anp_instalacoes.csv: {n_anp} linhas")
    print(f"bases_autorizadas.csv: {n_bases} linhas")
    print(f"empresas_alias.csv: {n_alias} linhas")


if __name__ == "__main__":
    main()
